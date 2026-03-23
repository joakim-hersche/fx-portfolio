"""GARCH model Excel report generator.

Single public function: export_garch_report — returns in-memory xlsx bytes.
"""

import io
import math
from datetime import datetime

import numpy as np
import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── Colour palette ────────────────────────────────────────────────────────────

_GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE") if _HAS_OPENPYXL else None  # type: ignore[possibly-undefined]
_AMBER_FILL  = PatternFill("solid", fgColor="FFEB9C") if _HAS_OPENPYXL else None  # type: ignore[possibly-undefined]
_RED_FILL    = PatternFill("solid", fgColor="FFC7CE") if _HAS_OPENPYXL else None  # type: ignore[possibly-undefined]
_HEADER_FILL = PatternFill("solid", fgColor="1E293B") if _HAS_OPENPYXL else None  # type: ignore[possibly-undefined]


def _bold_header(ws, row: int, cols: list) -> None:
    """Write a header row with white bold text on dark background."""
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")  # type: ignore[possibly-undefined]
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")  # type: ignore[possibly-undefined]


def _auto_width(ws) -> None:
    """Set column widths based on maximum content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)  # type: ignore[possibly-undefined]
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 45)


def _fmt_val(v, decimals: int = 4) -> object:
    """Return a rounded float or the original value for non-numeric types."""
    if v is None:
        return "—"
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return "∞" if math.isinf(v) else "—"
    if isinstance(v, float):
        return round(v, decimals)
    return v


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb: "Workbook", portfolio: dict, price_data: dict,
                   garch_params: dict, mc_result: dict, currency: str) -> None:
    ws = wb.create_sheet("Summary")
    _bold_header(ws, 1, ["Ticker", "Calibration window (days)", "Model", "Low confidence"])

    tickers = mc_result.get("tickers_used", list(garch_params.keys()))
    for r, ticker in enumerate(tickers, start=2):
        gp = garch_params.get(ticker, {})
        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=mc_result.get("train_days", "—"))
        ws.cell(row=r, column=3, value=gp.get("model", "—"))
        ws.cell(row=r, column=4, value="Yes" if gp.get("low_confidence") else "No")

    footer_row = len(tickers) + 3
    ws.cell(row=footer_row, column=1, value="Date generated:")
    ws.cell(row=footer_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws.cell(row=footer_row + 1, column=1, value="Currency:")
    ws.cell(row=footer_row + 1, column=2, value=currency)

    _auto_width(ws)


def _sheet_garch_params(wb: "Workbook", garch_params: dict,
                        model_comparison: dict) -> None:
    ws = wb.create_sheet("GARCH Parameters")
    headers = [
        "Ticker", "ω", "α", "β", "ν (df)",
        "Long-run vol (ann.)", "Half-life (days)", "Persistence (α+β)",
        "Converged", "Model",
    ]
    _bold_header(ws, 1, headers)

    for r, (ticker, gp) in enumerate(garch_params.items(), start=2):
        lr_var = gp.get("long_run_var")
        if lr_var is not None:
            lr_vol_ann = f"{math.sqrt(lr_var) * math.sqrt(252) / 100:.2%}"
        else:
            lr_vol_ann = "non-stationary"

        hl = gp.get("half_life")
        hl_str = f"~{hl:.0f} days" if hl is not None else "non-stationary"

        nu = gp.get("nu", np.inf)
        nu_str = "∞" if (nu is None or (isinstance(nu, float) and math.isinf(nu))) else round(nu, 2)

        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=_fmt_val(gp.get("omega"), 6))
        ws.cell(row=r, column=3, value=_fmt_val(gp.get("alpha"), 4))
        ws.cell(row=r, column=4, value=_fmt_val(gp.get("beta"), 4))
        ws.cell(row=r, column=5, value=nu_str)
        ws.cell(row=r, column=6, value=lr_vol_ann)
        ws.cell(row=r, column=7, value=hl_str)
        ws.cell(row=r, column=8, value=_fmt_val(gp.get("persistence"), 4))
        ws.cell(row=r, column=9, value="Yes" if gp.get("converged") else "No")
        ws.cell(row=r, column=10, value=gp.get("model", "—"))

        # Conditional formatting
        mc = model_comparison.get(ticker, {})
        preferred = mc.get("preferred", "comparable")
        converged = gp.get("converged", False)
        if converged and preferred == "garch-t":
            fill = _GREEN_FILL
        elif converged and preferred != "garch-t":
            fill = _AMBER_FILL
        else:
            fill = _RED_FILL
        for col in range(1, 11):
            ws.cell(row=r, column=col).fill = fill

    _auto_width(ws)


def _sheet_model_comparison(wb: "Workbook", garch_params: dict,
                             model_comparison: dict) -> None:
    ws = wb.create_sheet("Model Comparison")
    headers = [
        "Ticker", "GARCH-t LL", "GARCH-t AIC",
        "Constant-vol LL", "Constant-vol AIC",
        "ΔAIC", "Preferred", "Interpretation",
    ]
    _bold_header(ws, 1, headers)

    for r, ticker in enumerate(garch_params.keys(), start=2):
        mc = model_comparison.get(ticker, {})
        delta = mc.get("delta_aic", 0.0)
        preferred = mc.get("preferred", "comparable")

        if delta > 4:
            interp = "GARCH-t fits significantly better (ΔAIC > 4)"
        elif delta < -4:
            interp = "Constant-vol is adequate for this ticker (ΔAIC < -4)"
        else:
            interp = "Models are comparable (|ΔAIC| ≤ 4)"

        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=_fmt_val(mc.get("garch_ll"), 2))
        ws.cell(row=r, column=3, value=_fmt_val(mc.get("garch_aic"), 2))
        ws.cell(row=r, column=4, value=_fmt_val(mc.get("constant_ll"), 2))
        ws.cell(row=r, column=5, value=_fmt_val(mc.get("constant_aic"), 2))
        ws.cell(row=r, column=6, value=_fmt_val(delta, 2))
        ws.cell(row=r, column=7, value=preferred)
        ws.cell(row=r, column=8, value=interp)

    _auto_width(ws)


def _sheet_residual_diagnostics(wb: "Workbook", diagnostics: dict) -> None:
    ws = wb.create_sheet("Residual Diagnostics")
    headers = [
        "Ticker",
        "JB stat", "JB p-value", "JB pass",
        "LB stat (resid)", "LB p (resid)", "LB pass (resid)",
        "LB stat (resid²)", "LB p (resid²)", "LB pass (resid²)",
        "Residual kurtosis", "Residual skewness",
    ]
    _bold_header(ws, 1, headers)

    for r, (ticker, d) in enumerate(diagnostics.items(), start=2):
        gp = d.get("garch_params", {})
        sr_series = None  # residuals are removed from diagnostics dict for size

        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=_fmt_val(d.get("jb_stat"), 2))
        ws.cell(row=r, column=3, value=_fmt_val(d.get("jb_pvalue"), 4))
        ws.cell(row=r, column=4, value="Pass" if d.get("jb_normal") else "Fail")
        ws.cell(row=r, column=5, value=_fmt_val(d.get("lb_stat"), 2))
        ws.cell(row=r, column=6, value=_fmt_val(d.get("lb_pvalue"), 4))
        ws.cell(row=r, column=7, value="Pass" if d.get("lb_independent") else "Fail")
        ws.cell(row=r, column=8, value=_fmt_val(d.get("lb_sq_stat"), 2))
        ws.cell(row=r, column=9, value=_fmt_val(d.get("lb_sq_pvalue"), 4))
        ws.cell(row=r, column=10, value="Pass" if d.get("lb_sq_pass") else "Fail")
        # kurtosis/skewness stored in ticker_flags or computed separately
        ws.cell(row=r, column=11, value="—")
        ws.cell(row=r, column=12, value="—")

    _auto_width(ws)


def _sheet_correlation_matrix(wb: "Workbook", mc_result: dict,
                               garch_params: dict) -> None:
    ws = wb.create_sheet("Correlation Matrix")

    corr_df = mc_result.get("correlation_matrix")

    if corr_df is None or (isinstance(corr_df, pd.DataFrame) and corr_df.empty):
        ws.cell(row=1, column=1, value="Correlation matrix not available (single ticker or insufficient overlap).")
        return

    tickers = list(corr_df.columns)
    # Header row
    ws.cell(row=1, column=1, value="")
    for c, t in enumerate(tickers, start=2):
        cell = ws.cell(row=1, column=c, value=t)
        cell.font = Font(bold=True)  # type: ignore[possibly-undefined]

    for r, row_ticker in enumerate(tickers, start=2):
        ws.cell(row=r, column=1, value=row_ticker).font = Font(bold=True)  # type: ignore[possibly-undefined]
        for c, col_ticker in enumerate(tickers, start=2):
            val = corr_df.loc[row_ticker, col_ticker]
            cell = ws.cell(row=r, column=c, value=round(float(val), 4))
            if row_ticker == col_ticker:
                cell.fill = PatternFill("solid", fgColor="DBEAFE")  # type: ignore[possibly-undefined]

    _auto_width(ws)


def _sheet_simulation_summary(wb: "Workbook", mc_result: dict,
                               currency: str) -> None:
    ws = wb.create_sheet("Simulation Summary")

    horizon_map = {"3 months (63d)": 62, "6 months (126d)": 125, "1 year (252d)": 251}
    paths = mc_result.get("portfolio_paths")
    start_value = mc_result.get("start_value", 0.0)

    sym = currency

    metrics = [
        "Median outcome",
        "p10",
        "p90",
        "VaR 95%",
        "CVaR 95%",
        "Prob. above current",
    ]

    _bold_header(ws, 1, ["Metric"] + list(horizon_map.keys()))

    for r, metric in enumerate(metrics, start=2):
        ws.cell(row=r, column=1, value=metric)

    if paths is not None:
        for col, (label, day_idx) in enumerate(horizon_map.items(), start=2):
            idx = min(day_idx, paths.shape[1] - 1)
            end_vals = paths[:, idx]

            p50 = float(np.percentile(end_vals, 50))
            p10 = float(np.percentile(end_vals, 10))
            p90 = float(np.percentile(end_vals, 90))

            returns = (end_vals - start_value) / start_value if start_value > 0 else end_vals * 0
            var_pct = float(-np.percentile(returns, 5))
            tail_mask = returns <= -var_pct
            cvar_pct = float(-returns[tail_mask].mean()) if tail_mask.any() else var_pct
            prob_above = float((end_vals >= start_value).mean() * 100)

            row_vals = [
                f"{sym} {p50:,.0f}",
                f"{sym} {p10:,.0f}",
                f"{sym} {p90:,.0f}",
                f"{var_pct * 100:.1f}%",
                f"{cvar_pct * 100:.1f}%",
                f"{prob_above:.1f}%",
            ]
            for r_offset, val in enumerate(row_vals):
                ws.cell(row=2 + r_offset, column=col, value=val)
    else:
        for col in range(2, len(horizon_map) + 2):
            for r in range(2, len(metrics) + 2):
                ws.cell(row=r, column=col, value="—")

    _auto_width(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def export_garch_report(
    portfolio: dict,
    price_data: dict,
    garch_params: dict,
    model_comparison: dict,
    mc_result: dict,
    diagnostics: dict,
    currency: str,
) -> bytes:
    """
    Generate a GARCH model Excel report and return it as in-memory bytes.

    Sheets:
      1. Summary
      2. GARCH Parameters
      3. Model Comparison
      4. Residual Diagnostics
      5. Correlation Matrix
      6. Simulation Summary

    Parameters
    ----------
    portfolio : dict
    price_data : dict
    garch_params : dict   — {ticker: param_dict} from run_monte_carlo_portfolio
    model_comparison : dict — {ticker: comparison_dict}
    mc_result : dict      — full return from run_monte_carlo_portfolio
    diagnostics : dict    — from compute_model_diagnostics
    currency : str

    Returns
    -------
    bytes — xlsx file content
    """
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    wb = Workbook()  # type: ignore[possibly-undefined]
    # Remove the default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_summary(wb, portfolio, price_data, garch_params, mc_result, currency)
    _sheet_garch_params(wb, garch_params, model_comparison)
    _sheet_model_comparison(wb, garch_params, model_comparison)
    _sheet_residual_diagnostics(wb, diagnostics)
    _sheet_correlation_matrix(wb, mc_result, garch_params)
    _sheet_simulation_summary(wb, mc_result, currency)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
